import pandas as pd
import os
from openpyxl import load_workbook
import csv



def load_data(filename):
    return pd.read_csv(filename, low_memory=False)

def save_to_csv(df, filename):
	df.to_csv(filename, index=False)

if __name__ == "__main__":

	state_code = {
	    "AL": "Alabama",
	    "AK": "Alaska",
	    "AS": "American Samoa",
	    "AZ": "Arizona",
	    "AR": "Arkansas",
	    "CA": "California",
	    "CO": "Colorado",
	    "CT": "Connecticut",
	    "DE": "Delaware",
	    "DC": "District Of Columbia",
	    "FM": "Federated States Of Micronesia",
	    "FL": "Florida",
	    "GA": "Georgia",
	    "GU": "Guam",
	    "HI": "Hawaii",
	    "ID": "Idaho",
	    "IL": "Illinois",
	    "IN": "Indiana",
	    "IA": "Iowa",
	    "KS": "Kansas",
	    "KY": "Kentucky",
	    "LA": "Louisiana",
	    "ME": "Maine",
	    "MH": "Marshall Islands",
	    "MD": "Maryland",
	    "MA": "Massachusetts",
	    "MI": "Michigan",
	    "MN": "Minnesota",
	    "MS": "Mississippi",
	    "MO": "Missouri",
	    "MT": "Montana",
	    "NE": "Nebraska",
	    "NV": "Nevada",
	    "NH": "New Hampshire",
	    "NJ": "New Jersey",
	    "NM": "New Mexico",
	    "NY": "New York",
	    "NC": "North Carolina",
	    "ND": "North Dakota",
	    "MP": "Northern Mariana Islands",
	    "OH": "Ohio",
	    "OK": "Oklahoma",
	    "OR": "Oregon",
	    "PW": "Palau",
	    "PA": "Pennsylvania",
	    "PR": "Puerto Rico",
	    "RI": "Rhode Island",
	    "SC": "South Carolina",
	    "SD": "South Dakota",
	    "TN": "Tennessee",
	    "TX": "Texas",
	    "UT": "Utah",
	    "VT": "Vermont",
	    "VI": "Virgin Islands",
	    "VA": "Virginia",
	    "WA": "Washington",
	    "WV": "West Virginia",
	    "WI": "Wisconsin",
	    "WY": "Wyoming"
	}
	
	directory = "datasets/state_commodity_export/"
	print ("Loading data from file..")

	wb2 = load_workbook(directory + 'state_detail_by_commodity_cy.xlsx')

	allData = pd.DataFrame()

	for sheet in wb2.sheetnames:
		if sheet in set(['Commodity key', 'US total agricultural exports', 'US']):
			continue

		print (sheet)
		sh = wb2[sheet]
		sh['A2'] = "commodity"

		csvFile = open(directory + sheet + '.csv', 'w', newline="")
		c = csv.writer(csvFile)
		for index, r in enumerate(sh.rows):
			if index in set([0, 2, 3]) or index >= 28:
				continue
			c.writerow([cell.value for cell in r])
		
		csvFile.close()

		data = load_data(directory + sheet + '.csv')
		data.insert(1, "state", state_code[sheet])
		os.remove(directory + sheet + '.csv')

		# save_to_csv(data, directory + sheet + '.csv')
		allData = allData.append(data)
        
allData['commodity'] = allData['commodity'].str.replace(' [0-9]/', '', regex = True)
save_to_csv(allData, 'datasets/state_food_exports.csv')